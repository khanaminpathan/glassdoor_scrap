from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response

from bs4 import BeautifulSoup
import requests
import openpyxl
from urllib.request import Request, urlopen

# Create your views here.


class ScrapView(APIView):
    def get(self, request, format=None):
        excel=openpyxl.Workbook()
        # print(excel.sheetnames)
        sheet=excel.active
        sheet.title='Glassdoor Job Detail'
        # print(excel.sheetnames)
        sheet.append(['Company Name', 'Job Title', 'Salary Detail', 'URL'])


        # https://www.glassdoor.fr/Salaires/paris-digital-salaire-SRCH_IL.0,5_IM1080_KO6,13.htm?clickSource=searchBtn
        # https://www.glassdoor.fr/Salaires/paris-digital-salaire-SRCH_IL.0,5_IM1080_KO6,13_IP2.htm?clickSource=searchBtn

        hdr = {'User-Agent': 'Mozilla/5.0'}
        # all_pages=list(range(1, 6))
        # import pdb;
        # pdb.set_trace()
        # for all_pages in all_pages:
        # location=input("Enter City Name: ").lower()
        # jobtitle=input("Enter Job Title Name: ")
        jobtitle=""
        # location=""
        # print(jobtitle)
        # import pdb; pdb.set_trace()
        if 'jobtitle' in request.GET:
                if request.GET['jobtitle']:
                    jobtitle=request.GET['jobtitle']
                    jobtitle=jobtitle.replace(" ", "-").lower()
        count=len(jobtitle)
        count=count+6
        print(count)
        # if 'location' in request.GET:
        #         if request.GET['location']:
        #             location=request.GET['location']
        #             location=location.lower()
        # if 'fullurl' in request.GET:
        #     if request.GET['fullurl']:
        #         fullurl=request.GET['fullurl']
        data = []
        # var = 6
        url = 'https://www.glassdoor.fr/Salaires/paris-'+jobtitle + '-salaire-SRCH_IL.0,5_IM1080_KO6,'+ str(count)+'.htm?clickSource=searchBtn'
        source = Request(url.format(1), headers=hdr)
        page = urlopen(source)
        # print(len(page.read()))
        soup = BeautifulSoup(page, 'html.parser')
        jobs = soup.find_all('a', class_="css-f3vw95 e1aj7ssy3")
        description = soup.find_all('div', class_="col-12 col-lg px-xsm")
        price = soup.find_all("h3", class_="m-0 css-g261rn")
        url = soup.find_all('a', class_="css-f3vw95 e1aj7ssy3", href=True)
        ab = "https://www.glassdoor.fr"
        page_ul=soup.find('ul',class_="css-112ut70 e13qs2070")
        if page_ul:
            y = page_ul.find_all('li')
        else:
            return Response({'status': 0,
            'message': 'Data Not Found'} )

        # print(len(y))
        var = len(y)
        # page_ul
        l = []
        m = []
        n = []
        o = []
        result = []
        for item in jobs:
            l.append(item.text)
        for item in description:
            m.append(item.text)
        for item in price:
            if "€" in item.text:
                n.append(item.text)
        for url in url:
            o.append(ab+url['href'])
        # data.extend(l)
        try:
            for i in range(len(l)):
                
                data.append([l[i],m[i],n[i],o[i]])
                # break
        except Exception as e:
            print(e)
            # writer = csv.writer(f)
            # row="Title,Description,Salary"
            # writer.writerow(row)
                    
                    # print(l[i],m[i],n[i],o[i])
                    # break
                    # sheet.append([l[i]])
        # sheet.append([l[i],m[i],n[i],o[i]])
        sheet.append(data[0])
        for all_pages in range(2,var+1):
            # source = Request('https://www.glassdoor.fr/Salaires/france-digital-marketing-salaire-SRCH_IL.0,6_IN86_KO7,24.htm?clickSource=searchBtn',headers=hdr)
            # if all_pages == 1:
            #     url = 'https://www.glassdoor.fr/Salaires/paris-'+jobtitle + '-salaire-SRCH_IL.0,5_IM1080_KO6,'+ str(count)+'.htm?clickSource=searchBtn'
            #     print(url)
            #     source = Request(url.format(all_pages), headers=hdr)
            # else:
            url = 'https://www.glassdoor.fr/Salaires/paris-'+jobtitle + '-salaire-SRCH_IL.0,5_IM1080_KO6,'+ str(count) +'_IP'+ str(all_pages)+'.htm?clickSource=searchBtn'
            print(url)
            source = Request(url.format(all_pages), headers=hdr)
            # if all_pages == 1:
            #     url = fullurl
            #     print(url)
            #     source = Request(url.format(all_pages), headers=hdr)
            # else:
            #     url = fullurl
            #     print(url)
            #     source = Request(url.format(all_pages), headers=hdr)
            # print(url)
            # print(source)
            page = urlopen(source)
            # print(len(page.read()))
            soup = BeautifulSoup(page, 'html.parser')
            jobs = soup.find_all('a', class_="css-f3vw95 e1aj7ssy3")
            description = soup.find_all('div', class_="col-12 col-lg px-xsm")
            price = soup.find_all("h3", class_="m-0 css-g261rn")
            url = soup.find_all('a', class_="css-f3vw95 e1aj7ssy3", href=True)
            ab = "https://www.glassdoor.fr"
            page_ul=soup.find('ul',class_="css-112ut70 e13qs2070")
            # y = page_ul.find_all('li')
            # print(len(y))
            # var = len(y)
            # page_ul
            l = []
            m = []
            n = []
            o = []
            result = []
            for item in jobs:
                l.append(item.text)
            for item in description:
                m.append(item.text)
            for item in price:
                if "€" in item.text:
                    n.append(item.text)
            for url in url:
                o.append(ab+url['href'])
            # data.extend(l)
            try:
                for i in range(len(l)):
                    
                    data.append([l[i],m[i],n[i],o[i]])
                    # break
            except Exception as e:
                print(e)
            # writer = csv.writer(f)
            # row="Title,Description,Salary"
            # writer.writerow(row)
                    
                    # print(l[i],m[i],n[i],o[i])
                    # break
                    # sheet.append([l[i]])
        # sheet.append([l[i],m[i],n[i],o[i]])
        sheet.append(data[0])
            
                    # writer.writerow(l[i]+","+m[i]+","+str(n[i]))
                # f.close()
            
        # print(len(data))
        excel.save('Glassdoor New Salary Detail.xlsx')
        return Response({'status': 1,
            'message': 'Success',
            'data': data} )
        # return Response({'status': 1,
        #     'message': 'Success',
        #     'data': data[0]} )
        